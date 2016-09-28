# -*- coding: utf-8 -*-
# this file is released under public domain and you can use without limitations
#
#########################################################################
## This is a sample controller
## - index is the default action of any application
## - user is required for authentication and authorization
## - download is for downloading files uploaded in the db (does streaming)
####rowcontent['mobile'] in db().select(db.history_order.mobile)#########
# #######################################################################
import os
import re
from simhash import Simhash,SimhashIndex
#from snownlp import SnowNLP
import time
import hashlib
import openpyxl
import datetime
import redis
import xml.etree.ElementTree as XML_ET
import json
import requests

auth.settings.registration_requires_approval = False
auth.settings.registration_requires_verification = True
auth.settings.allow_basic_login = True
auth.define_tables(username=True)
auth.settings.login_next = URL('select_province')

rds=redis.StrictRedis()
today=datetime.date.today().strftime("%d")
yesterday=(datetime.date.today() + datetime.timedelta(days=-1)).strftime("%d")
auth.settings.allow_basic_login = True
auth.define_tables(username=True)
pat_dat=re.compile("[0-9]+")
provinces=(u'北京',u'天津',u'河北',u'山西',u'内蒙古',u'辽宁',u'吉林',u'黑龙江',u'上海',u'江苏',
           u'浙江',u'安徽',u'福建',u'江西',u'山东',u'河南',u'湖北',u'湖南',u'广东',u'广西',
           u'海南',u'重庆',u'四川',u'贵州',u'云南',u'西藏',u'陕西',u'甘肃',u'青海',u'宁夏',
           u'新疆',u'香港',u'澳门',u'台湾')
zhixia=(u'北京',u'天津',u'重庆',u'上海')
texing=(u'香港',u'澳门')
SKU_dict={u'589203710279202':u'18片S码',
u'690314820783301':u'18片NB码',
u'589203710279204':u'8片S码',
u'690314820785503':u'8片NB码'}
top4_city=(u'北京市',u'成都市',u'广州市',u'上海市')
A_city=(u'合肥市',u'福州市',u'兰州市',u'深圳市','南宁市',u'贵阳市',u'海口市',u'石家庄市',
u'郑州市',u'哈尔滨市',u'武汉市',u'长沙市',u'南京市',u'大连市',u'沈阳市',u'呼和浩特市',
u'银川市',u'西宁市',u'济南市',u'青岛市',u'太原市',u'西安市',u'天津市',u'拉萨市',u'昆明市',
u'杭州市',u'重庆市')

def init_db():
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'订单编号',field_id='order_id',field_order=1)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'店铺名称',field_id='shop_name',field_order=2)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'外部平台单号',field_id='ex_id',field_order=3)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'电话',field_id='telephone',field_order=4)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'手机',field_id='mobile',field_order=5)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'收货人',field_id='user_name',field_order=6)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'收货省',field_id='province',field_order=7)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'收货市',field_id='city',field_order=8)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'收货县',field_id='county',field_order=9)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'收货地址',field_id='address',field_order=10)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'产品编号',field_id='product_id',field_order=11)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'条形码',field_id='barcode',field_order=12)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'规格',field_id='specification',field_order=13)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'网店品名',field_id='ex_product_name',field_order=14)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'产品名称',field_id='product_name',field_order=15)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'订货数量',field_id='order_num',field_order=16)
    db.db_map.insert(table_type='excel',table_name='custom_order',field_name=u'网店规格',field_id='ex_spec',field_order=17)

def d3tree():
    return dict()
def circle():
    return dict()
def force():
    return dict()

def import_csv(csvfile):
    db.custom_order.import_from_csv_file(csvfile)

def get_feature(s):
    return [s[i:i+1] for i in range(max(len(s)-1+1,1))]

def GetTimeStamp():
    t=time.localtime()
    lst=[]
    for i in xrange(len(t)):
        if i>=5:
            break
        v=str(t[i]).zfill(2)
        lst.append(v)
    reStr="".join(lst)
    return reStr

def log_file(user,msg):
    dir=os.getcwd()
    fileName="log_access.txt"
    finalPath=os.path.join(dir,fileName)
    with open(finalPath,"ab") as f:
        f.write("[%s:%s],%s\r\n"%(GetTimeStamp(),user,msg))
    return finalPath

def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/manage_users (requires membership in
    http://..../[app]/default/user/bulk_register
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    """
    form=auth()
    #form.insert(-1,TR(INPUT(_type="checkbox",_name="vehicle",_value="Bike"),"I have a bike",
    #INPUT(_type="checkbox",_name="vehicle",_value="Car"),"I have a car",BR()))
    return dict(form=form)

shop_ids=['108','109' ,'110','111','112','114','120','122',
         '23','29','36','38','61','63','64','93','94','95']
product_size=['NB12','S12','NB8','S8']
def census_form():
    cnt5={}
    s5=[]
    cnt6={}
    s6=[]
    for j in product_size:
        s=0
        for i in shop_ids:
            buf1=db((db.census_result.tstat=='5')&\
               (db.census_result.product_size==j)&\
               (db.census_result.shop_id==i)).select(
                db.census_result.census_count)
            cnt5[j+'_'+i]=buf1[0]['census_count']
            s=s+int(cnt5[j+'_'+i])
        s5.append(s)
        s=0
        for i in shop_ids:
            buf2=db((db.census_result.tstat=='6')&\
               (db.census_result.product_size==j)&\
               (db.census_result.shop_id==i)).select(
                db.census_result.census_count)
            cnt6[j+'_'+i]=buf2[0]['census_count']
            s=s+int(cnt6[j+'_'+i])
        s6.append(s)
    s6s=s6[0]+s6[1]+s6[2]+s6[3]
    s5s=s5[0]+s5[1]+s5[2]+s5[3]
    #form= FORM(
        #INPUT(_name='name',_type='date',requires=IS_NOT_EMPTY()),
        #INPUT(_type='submit'))
    return dict(
        #form=form,
        s5=s5,
        s5s=s5s,
        s6=s6,
        s6s=s6s,
        cnt5=cnt5,
        cnt6=cnt6,
        shop_ids=shop_ids)

@auth.requires_login()
def trace_trade():
    qry=''
    if (request.vars.begin_time) and (request.vars.end_time):
        begintime=request.vars.begin_time
        endtime=request.vars.end_time
        qry='(db.trade.order_date>"%s")& (db.trade.order_date<"%s")'%(begintime,endtime)
    shop_ids=[]
    if request.vars.shopids:
        t=request.vars.shopids
        if isinstance(t,str):
            shop_ids.append(t)
        else:
            for i in t:
                shop_ids.append(i)
    if shop_ids:
        qry1=''
        for i in shop_ids:
            qry1=qry1+'|'+'('+'db.trade.shop_id.like(u"%s")'%(i)+')'
        if qry1:
            qry='('+qry1[1:]+')'+'&'+qry
    if qry:
        query= eval_in_global_env(qry)
        search_result=db(query).select(db.trade.id,db.trade.mobilPhone,limitby=(0,100))
    else:
        search_result=""

    return dict(
        search_result=search_result,
        current_user='%(first_name)s' %(auth.user),
    )

def back_trade():
    return dict()

@auth.requires_login()
def search_by_phone():
    return dict(
        current_user='%(first_name)s' %(auth.user),
        log_state=0,
        can_log=1
    )
def search_trade():
    mobil=request.vars.input_mobil
    current_user=request.vars.user
    search_result = db(db.trade.mobilPhone.like(mobil)).select(db.trade.out_tid,db.trade.wrong_reason,db.trade.address,db.trade.status,db.trade.shop_id)
    log_file(current_user,mobil)
    if search_result:
        result=search_result.first()
        shopid=result['shop_id']
        recv_addr=result['address']
        if result['status']=='2':
            reason=result['wrong_reason']
            return DIV(DIV("shop_id:",shopid),BR(),
                       DIV("wrong_reason:",reason),BR(),
                       DIV("address:",recv_addr),
                       _align="left",_style="padding-left: 5cm")
        elif result['status']=='3':
            trackingno=db(db.WuLiuInfo.orderId.like(result['out_tid'])).select(db.WuLiuInfo.trackingNo)
            return DIV(DIV("shop_id:",shopid),BR(),
                       DIV("WuLiu:",trackingno.first()['trackingNo']),BR(),
                       DIV("address:",recv_addr),
                       _align="left",_style="padding-left: 5cm")
        else:
            return DIV(DIV("shop_id:",shopid),BR(),
                       DIV("订单未审核"),BR(),
                       DIV("address:",recv_addr),
                       _align="left",_style="padding-left: 5cm")
    else:
        return "此号码不存在"

def test_table():
    response.flash = "wang,wang,wang"
    return 'hello,hello,hello'

def test_flash():
    response.flash = "qqqqqqqqqq"
    return dict()

def EditTable():
    return dict()

def test_ajax():
    return dict()

def echo():
    name_dict=dict(request.vars)
    name_key=name_dict.keys()[0]
    name_val=name_dict.values()[0]
    response.flash = name_val
    #return request.vars.your_mes
    #return request.vars.name
    #return "jQuery('#target').html(%s);" % repr(name_val)
    #return "jQuery('#{{=targetid}}').html(%s);" % repr(request.vars.name1)
    #response.flash = 'please fill out the form'
    #return
    return "jQuery('#%s').css('color','red');"%(name_key)

def main():
    return log_file("wang","18794757802")

#@auth.requires_login()
def display_form():
    grid = SQLFORM.grid(db.wrong_order,fields=\
        [db.wrong_order.order_id,db.wrong_order.ex_id,db.wrong_order.mobile,
        db.wrong_order.user_name,db.wrong_order.province,db.wrong_order.city,
        db.wrong_order.address,db.wrong_order.wrong_reason,db.wrong_order.dup_ID,
        db.wrong_order.dup_ex,db.wrong_order.dup_address,db.wrong_order.dup_name,
        db.wrong_order.dup_phone])
    return locals()

def csv():
    import gluon.contenttype
    response.headers['Content-Type'] = \
        gluon.contenttype.contenttype('.csv')
    db = eval_in_global_env(request.args[0])
    if (request.vars.qryfield)and(request.vars.qryoperater)and(request.vars.qryvalue):
        qry=request.vars.qryfield+request.vars.qryoperater+request.vars.qryvalue
    else:
        qry = 'db.trade.id>0'
    query = eval_in_global_env(qry)
    response.headers['Content-disposition'] = 'attachment; filename=%s_%s_%s.csv'\
    % ('trade',request.vars.begin_line,request.vars.end_line)
    return str(db(query, ignore_common_filters=True).select(limitby=(int(request.vars.begin_line),int(request.vars.end_line))))
##################################################################################################################################

import copy,datetime
global_env = copy.copy(globals())
global_env['datetime'] = datetime
def eval_in_global_env(text):
    exec ('_ret=%s' % text, {}, global_env)
    return global_env['_ret']

def himin_distance(i1,i2):
    f=64
    x = (i1 ^ i2) & ((1 << 64) - 1)
    ans = 0
    while x:
        ans += 1
        x &= x - 1
    return ans

prov_dict={
    "bj":u'北京' ,"sh":u'上海' ,"tj":u'天津', "cq":u'重庆',
    "heb":u'河北', "sx":u'山西', "ln":u'辽宁', "jn":u'吉林',
    "js":u'江苏',"zj":u'浙江', "ah":u'安徽', "fj":u'福建' ,
    "jx":u'江西' ,"sd":u'山东' ,"hen":u'河南', "hub":u'湖北',
    "hun":u'湖南',"gd":u'广东' ,"gx":u'广西' ,"hn":u'海南' ,
    "sc":u'四川' ,"gz":u'贵州', "yn":u'云南',  "xz":u'西藏' ,
    "sax":u'陕西' ,"gs":u'甘肃' ,"qh":u'青海' ,"nx":u'宁夏',
    "xj":u'新疆', "hlj":u'黑龙江',"nm":u'内蒙'}

@auth.requires_login()
def select_province():
    if request.vars.refresh:
        for i in rds.keys('Pvc_*'):
            rds.delete(i)
        for i in  rds.keys('otid*'):
            rds.delete(i)
    db = eval_in_global_env('db')
    current_user='%(first_name)s' %(auth.user)
    prov_key=["bj","sh" ,"tj", "cq",
              "ln", "jn","sd","sx", #东北
              "sax" ,"gs","nx","qh",#西北
              "hen","heb","hun","hub",#华中
              "jx","js","zj", "ah",#华东
              "gd" ,"gx" ,"hn" ,"fj",#华南
              "yn","gz","sc","xz",#西南
              "xj","nm","hlj"]#华北
    #prov_value=prov_dict.values()
    pvc_user='Pvc_'+current_user
    if request.vars.provice:
        t=request.vars.provice
        if isinstance(t,str):
            rds.sadd(pvc_user,t)
        else:
            for i in t:
                rds.sadd(pvc_user,i)
        redirect(URL('test_display_form'))
    prov_set=set()
    us=rds.keys('Pvc_*')
    which_day=1
    if request.vars.which_day:
        which_day= int(request.vars.which_day)
    that_day= (datetime.date.today() - datetime.timedelta(days=which_day)).strftime("%Y%m%d")
    next_day= (datetime.date.today() - datetime.timedelta(days=which_day-1)).strftime("%Y%m%d")
    select_day= that_day[0:4]+u"年"+that_day[4:6]+u"月"+that_day[6:]+u"日"

    prov_cnt={}
    for i in prov_key:
        qry1='db.trade.province.like(u"%s%s")'%(prov_dict[i],'%')
        qry='('+qry1+')'+'&'+'(db.trade.status!="2") & (db.trade.status!="3") &(db.trade.order_date>"%s")& (db.trade.order_date<"%s")'%(that_day,next_day)
        query= eval_in_global_env(qry)
        prov_cnt[i]=db(query).count()
    for i in us:
        s1=rds.smembers(i)
        prov_set=prov_set|s1
    return dict(
        prov_cnt=prov_cnt,
        prov_key=prov_key,
        prov_dict=prov_dict,
        prov_set=prov_set,
        which_day=which_day,
        select_day=select_day,
    )

@auth.requires_login()
def test_display_form():
    current_user='%(first_name)s' %(auth.user)
    pvc_user='Pvc_'+current_user
    every_otid='otid'+current_user
    if not request.function=='select_province' and not rds.exists(pvc_user):
        redirect(URL('select_province'))
    db = eval_in_global_env('db')
    try:
        is_imap = db._uri.startswith("imap://")
    except (KeyError, AttributeError, TypeError):
        is_imap = False
    step = 100

    table = "trade"
    which_day=1
    if request.vars.which_day:
        which_day= int(request.vars.which_day)
    that_day= (datetime.date.today() - datetime.timedelta(days=which_day)).strftime("%Y%m%d")
    next_day= (datetime.date.today() - datetime.timedelta(days=which_day-1)).strftime("%Y%m%d")
    qry='(db.trade.status!="2") & (db.trade.status!="3") &(db.trade.order_date>"%s")& (db.trade.order_date<"%s")'%(that_day,next_day)
    select_day= that_day[0:4]+u"年"+that_day[4:6]+u"月"+that_day[6:]+u"日"
    provs=rds.smembers(pvc_user)
    provS=[]
    if provs:
        qry1=''
        for i in provs:
            qry1=qry1+'|'+'('+'db.trade.province.like(u"%s%s")'%(prov_dict[i],'%')+')'
            provS.append(prov_dict[i])
        if qry1:
            qry='('+qry1[1:]+')'+'&'+qry
    query= eval_in_global_env(qry)
    #exec 'query=%s'%(qry)
    nrows = db(query).count()
    start=0
    if request.vars.start:
        start = int(request.vars.start)
    elif request.vars.locate:
        if 0< int(request.vars.locate) < nrows/step:
            start = (int(request.vars.locate)-1)*step
        else:
            response.flash = T("out of range")
    else:
        start = 0
    if start<nrows-step:
        stop = start + step
    else:
        stop=nrows
    rows=db(query).select(limitby=(start,stop),orderby=db.trade.out_tid)
    #nrows = db(query, ignore_common_filters=True).count()
    #rows = db(query, ignore_common_filters=True).select(limitby=(start,stop))
    tb=None
    simaddr=[]
    for r in rows:
        sim_buf=[]
        if r['simid']:
            otids=r['simid'].split('&')
            for i in otids:
                ar=db(db.history_order.id==i).select(db.history_order.id,
                    db.history_order.consignee,db.history_order.address,db.history_order.data_date).first()
                sim_buf.append(ar)
        simaddr.append(sim_buf)
        rds.sadd(every_otid,r['out_tid'])

    locateform = FORM(TABLE(TR(
    TD(INPUT(_name='locate', _value=request.vars.locate,_style='width:45px;')),
    TD(INPUT(_type='submit', _value=T('Locate'),_style='width:60px;text-align:center;color:blue;')))),
    _action=URL(r=request,args=request.args),_style="padding:0;margin:0;")

    return dict(
        locateform=locateform,
        select_day=select_day,
        provS=provS,#
        table=table,
        start=start,
        step=step,
        nrows=nrows,
        rows=rows,
        which_day=which_day,
        simaddr=simaddr,
        current_user=current_user,
        tb=tb
    )

def update_state():
    name_dict=dict(request.vars)
    user_val=name_dict['update_rest']
    every_otid='otid'+user_val
    name_dict.pop('update_rest')
    name_key=pat_dat.findall(name_dict.keys()[0])[0]
    name_val=name_dict.values()[0]
    if name_val:
        if rds.sismember(every_otid,name_key):
            rds.srem(every_otid,name_key)
        if name_val=='TG':
            db(db.trade.out_tid==name_key).update(status='3')
            r=db(db.trade.out_tid==name_key).select(db.trade.status)
            if r:
                if r[0]['status']=='3':
                    response.flash = "trade has been updated"
                else:
                    response.flash = "failed to update the trade"
            else:
                response.flash = "failed to update the trade"
            return "jQuery('#%s').css('color','#00ff00');"%(name_key)
        else:
            db(db.trade.out_tid==name_key).update(**dict(status='2',wrong_reason=name_val))
            r=db(db.trade.out_tid==name_key).select(db.trade.status,db.trade.wrong_reason)
            if r:
                if r[0]['status']=='2' and r[0]['wrong_reason']==name_val:
                    response.flash = "trade has been updated"
                else:
                    response.flash = "failed to update the trade"
            else:
                response.flash = "failed to update the trade"
            return "jQuery('#%s').css('color','#ff00ff');"%(name_key)
    else:
        return ""

def update_all_rest():
    user_val=request.vars['update_rest']
    every_otid='otid'+user_val
    t=rds.smembers(every_otid)
    for i in t:
        r=db(db.trade.out_tid==i).update(status='3')
        if r:
            rds.srem(every_otid,i)
    if rds.smembers(every_otid):
        response.flash = "failed to update all rest"
    else:
        response.flash = "all rest has been updated"
    return ''

def EditTableAjax():
    name_dict=dict(request.vars)
    namekey=name_dict.keys()[0]
    name_key=pat_dat.findall(namekey)[0]
    name_val=name_dict.values()[0]
    r=db(db.trade.out_tid==name_key).update(address=name_val)
    if r:
        response.flash = "address have been updated"
    else:
        response.flash = "no updated"
    return ''

def add2blackname():
    name_dict=dict(request.vars)
    #namekey=name_dict.keys()[0]
    #name_key=pat_dat.findall(namekey)[0]
    name_val=name_dict.values()[0]
    r=rds.sadd('blackname',name_val)
    log_file("blackname",name_val)
    if r:
        response.flash = "add to blackname successfully"
    else:
        response.flash = "no added"
    return ''

def add2may_blackname():
    name_dict=dict(request.vars)
    name_val=name_dict.values()[0]
    r=rds.sadd('may_blackname',name_val)
    log_file("may_blackname",name_val)
    if r:
        response.flash = "add to may_blackname successfully"
    else:
        response.flash = "no added"
    return ''

def echo_rownum():
    trade_items=db(db.trade.mobilPhone.like(request.vars.id)).select().as_list()[0]
    return ''.join([DIV(SPAN(k,_class='items'),SPAN(v,_class='item')).xml() for k,v in trade_items if k])

def month_input():
    return dict()

def month_selector():
    if not request.vars.month:
        return ''
    months = ['January', 'February', 'March', 'April', 'May',
    'June', 'July', 'August', 'September' ,'October',
    'November', 'December']
    month_start = request.vars.month.capitalize()
    selected = [m for m in months if m.startswith(month_start)]
    return DIV(*[DIV(k,
    _onclick="jQuery('#month').val('%s')" % k,
    _onmouseover="this.style.backgroundColor='yellow'",
    _onmouseout="this.style.backgroundColor='white'"
    ) for k in selected])

def add_db():
    rows=db(db.history_order.id>0).select()
    r0=rows.first()
    #r0.out_tid=r0.tid
    r0.update_record(out_tid=r0.tid)
    return locals()

def list_items():
    items = db().select(db.db_map.ALL)
    return dict(items=items)

def vote():
    item = db.trade[request.vars.id]
    new_votes = item.votes + 1
    item.update_record(votes=new_votes)
    return str(new_votes)

def display_pampers_result():
    grid = SQLFORM.grid(db.pampers_result)
    return locals()

def display_pampers_order():
    grid = SQLFORM.grid(db.pampers_order)
    return locals()

def display_trade_order():
    grid = SQLFORM.grid(db.trade)
    return locals()

def check_order():
    orderList=db(db.WuLiuInfo).select().as_list()
    return locals()

def trade_proc():
    essential_field=['consignee','address','province','city','area','mobilPhone','out_tid','shop_id','order_date','barCode','product_title','standard','out__tid']
    trade_list=db(db.trade).select().as_list()
    trade_draw=[]
    for each_trade in trade_list[0:30]:
        content={}
        for field in essential_field:
            content[field]=each_trade[field]
        trade_draw.append(content)
    WuLiu_list=db(db.WuLiuInfo).select().as_list()
    WuLiu_draw=[]
    for i in range(0,len(WuLiu_list)/10):
        WuLiu_draw.append(WuLiu_list[i*10:(i+1)*10])
    WuLiu_draw.append(WuLiu_list[(len(WuLiu_list)/10)*10:len(WuLiu_list)])
    return locals()

def process_order():
    Info=request.vars
    order_id=Info["order_id"]
    iFlag=Info["process_flag"]

def data_wrong():
    form = SQLFORM(db.wrong_order)
    if form.process().accepted:
        response.flash = 'form accepted'
    elif form.errors:
        response.flash = 'form has errors'
    else:
        response.flash = 'please fill out the form'
    return dict(form=form)

def import_excel(file):
    WB=openpyxl.load_workbook(file)
    WS=WB.active
    return WS

def add_excel(ws):
    field_list = db(db.db_map.table_name=='custom_order').select(db.db_map.field_name,db.db_map.field_id).as_list()
    field_dict = {}
    for fd in field_list:
        field_dict[fd['field_name'].decode("utf8")]=fd['field_id']
    header=[] #表头
    for cols in range(1,ws.max_column+1):
        header.append(field_dict[ws.cell(row=1,column=cols).value])
    buf_list=[]
    for rows in range(467,ws.max_row+1):
        rowcontent={}
        for cols in range(1,ws.max_column+1):
            rowcontent[header[cols-1]]=ws.cell(row=rows, column=cols).value
        rowcontent['address']=rowcontent['address'].replace(rowcontent['province']+u'','')
        rowcontent['address']=rowcontent['address'].replace(u''+rowcontent['city'],'')
        rowcontent['address']=rowcontent['address'].replace(rowcontent['county']+u'','')
        if ws.cell(row=rows, column=9).value !=ws.cell(row=rows-1,column=9).value:
            buf_list=db((db.history_order.province==ws.cell(row=rows, column=7).value)&\
                        (db.history_order.city==ws.cell(row=rows, column=8).value)&\
                        (db.history_order.county==ws.cell(row=rows, column=9).value)&\
                        (db.history_order.product_name==SKU_dict[ws.cell(row=rows,column=12).value])).select\
                (db.history_order.mobile,db.history_order.address_bak,db.history_order.order_id,\
                 db.history_order.ex_id,db.history_order.user_name).as_list()
        is_correct=1
        for i in range(0,len(buf_list)):
            if rowcontent['mobile']==buf_list[i]['mobile']:
                is_correct=0
                rowcontent['wrong_reason']=u'与历史号码第'+str(buf_list[i]['order_id'])+u'条重复'
                rowcontent['dup_ID']=buf_list[i]['order_id']
                rowcontent['dup_ex']=buf_list[i]['ex_id']
                rowcontent['dup_address']=buf_list[i]['address_bak']
                rowcontent['dup_name']=buf_list[i]['user_name']
                rowcontent['dup_phone']=buf_list[i]['mobile']
                db.wrong_order.insert(**rowcontent)
                break
            elif Simhash(get_feature(buf_list[i]['address_bak'].decode('UTF-8'))).distance(Simhash(get_feature(u''+rowcontent['address'])))<18:
                is_correct=0
                rowcontent['wrong_reason']=u'与历史地址第'+str(buf_list[i]['order_id'])+u'条重复'
                rowcontent['dup_ID']=buf_list[i]['order_id']
                rowcontent['dup_ex']=buf_list[i]['ex_id']
                rowcontent['dup_address']=buf_list[i]['address_bak']
                rowcontent['dup_name']=buf_list[i]['user_name']
                rowcontent['dup_phone']=buf_list[i]['mobile']
                db.wrong_order.insert(**rowcontent)
                break
            else:continue
        if is_correct==1:
            if rowcontent['user_name']!= ws.cell(row=rows-1, column=6).value:
                rowcontent['product_tag']=SKU_dict[ws.cell(row=rows,column=12).value]
                db.custom_order.insert(**rowcontent)
    for rows in range(2,467):
        rowcontent={}
        for cols in range(1,6+1)+range(11,ws.max_column+1):
            rowcontent[header[cols-1]]=ws.cell(row=rows, column=cols).value
        for prov in provinces:
            if prov in str(ws.cell(row=rows,column=10).value):
                spare_addr=ws.cell(row=rows,column=10).value.replace(prov,'')
                spare_addr=spare_addr.replace(u'省','')
                spare_addr=spare_addr.replace(u'市','')
                if prov in zhixia:
                    rowcontent[header[6]]=prov+u'市'
                elif prov in texing:
                    rowcontent[header[6]]=prov+u'特别行政区'
                elif prov==u'广西':
                    rowcontent[header[6]]=prov+u'壮族自治区'
                elif prov==u'新疆':
                    rowcontent[header[6]]=prov+u'维吾尔自治区'
                elif prov==u'宁夏':
                    rowcontent[header[6]]=prov+u'回族自治区'
                elif prov==u'西藏':
                    rowcontent[header[6]]=prov+u'自治区'
                elif prov==u'内蒙古':
                    rowcontent[header[6]]=prov+u'自治区'
                else:
                    rowcontent[header[6]]=prov+u'省'
                break
            else:
                spare_addr=ws.cell(row=rows,column=10).value
        addr2=re.split(u'市|自治州',spare_addr)
        rowcontent[header[7]]=addr2[0]
        spare_addr2=re.sub(addr2[0],'',spare_addr)
        spare_addr2=re.sub(u'市|自治州','',spare_addr2)

        addr3=re.split(u'县|区',spare_addr2)
        rowcontent[header[8]]=addr3[0]+u'(县)'
        spare_addr3=re.sub(addr3[0],'',spare_addr2)
        spare_addr3=re.sub(u'县|区','',spare_addr3)
        rowcontent[header[9]]=spare_addr3
        if (header[6] not in rowcontent):
            db.address_lack.insert(**rowcontent)
        else:
            buf_list=db((rowcontent[header[6]]==db.history_order.province)&\
                        (db.history_order.product_name==SKU_dict[ws.cell(row=rows,column=12).value])).select\
                (db.history_order.mobile,db.history_order.address,db.history_order.order_id,\
                 db.history_order.ex_id,db.history_order.user_name).as_list()
            is_correct=1
            for i in range(0,len(buf_list)):
                if rowcontent['mobile']==buf_list[i]['mobile']:
                    is_correct=0
                    rowcontent['wrong_reason']=u'与历史号码第'+str(buf_list[i]['order_id'])+u'条重复'
                    rowcontent['dup_ID']=buf_list[i]['order_id']
                    rowcontent['dup_ex']=buf_list[i]['ex_id']
                    rowcontent['dup_address']=buf_list[i]['address']
                    rowcontent['dup_name']=buf_list[i]['user_name']
                    rowcontent['dup_phone']=buf_list[i]['mobile']
                    db.wrong_order.insert(**rowcontent)
                    break
                elif Simhash(get_feature(buf_list[i]['address'])).distance(Simhash(get_feature(ws.cell(row=rows,column=10).value)))<24:
                    is_correct=0
                    rowcontent['wrong_reason']=u'与历史地址第'+str(buf_list[i]['order_id'])+u'条重复'
                    rowcontent['dup_ID']=buf_list[i]['order_id']
                    rowcontent['dup_ex']=buf_list[i]['ex_id']
                    rowcontent['dup_address']=buf_list[i]['address']
                    rowcontent['dup_name']=buf_list[i]['user_name']
                    rowcontent['dup_phone']=buf_list[i]['mobile']
                    db.wrong_order.insert(**rowcontent)
                    break
                else:continue
            if is_correct==1:
                if rowcontent['user_name']!= ws.cell(row=rows-1, column=6).value:
                    rowcontent['product_tag']=SKU_dict[ws.cell(row=rows,column=12).value]
                    db.custom_order.insert(**rowcontent)

pampers_dict={'order_id':'A','ex_id':'B','user_name':'C','mobile':'D','province':'E',
              'city':'F','county':'G','address':'H','pregnancy':'I','product_piece':'J',
              'product_size':'K','datetime':'L'}
def add_pampers(ws1):
    header=['order_id','ex_id','user_name','mobile','province','city','county','address',
            'pregnancy','product_piece','product_size','date_time']
    for rows in range(650,ws1.max_row+1):
        rowcontent={}
        buf_list=[]
        for cols in range(1,ws1.min_col+1):
            rowcontent[header[cols-1]]=u''+str(ws1.cell(row=rows,column=cols).value)
        rowcontent['address']=rowcontent['address'].replace(rowcontent['province'],'')
        rowcontent['address']=rowcontent['address'].replace(rowcontent['city'],'')
        rowcontent['address']=rowcontent['address'].replace(rowcontent['county'],'')
        if ws1.cell(row=rows,column=7).value==ws1.cell(row=rows-1,column=7).value:
            db.pampers_self_dup(**rowcontent)
        elif not isinstance(rowcontent['address'],str):
            db.pampers_addr_lack.insert(**rowcontent)
        else:
            buf_list=db((db.history_order.province==ws1.cell(row=rows, column=5).value)&\
                        (db.history_order.city==ws1.cell(row=rows, column=6).value)&\
                        (db.history_order.county==ws1.cell(row=rows, column=7).value)).select\
                (db.history_order.mobile,db.history_order.address_bak,db.history_order.order_id,
                 db.history_order.ex_id,db.history_order.user_name,db.history_order.product_name).as_list()
        is_correct=1
        for i in range(0,len(buf_list)):
            if (buf_list[i]['product_name'].decode('utf8')).startswith(str(rowcontent['product_piece'])):
                if rowcontent['mobile']==buf_list[i]['mobile']:
                    is_correct=0
                    rowcontent['wrong_reason']=u'与历史号码第'+str(buf_list[i]['order_id'])+u'条重复'
                    rowcontent['dup_ID']=buf_list[i]['order_id']
                    rowcontent['dup_ex']=buf_list[i]['ex_id']
                    rowcontent['dup_address']=buf_list[i]['address_bak']
                    rowcontent['dup_name']=buf_list[i]['user_name']
                    rowcontent['dup_phone']=buf_list[i]['mobile']
                    db.pampers_history_dup.insert(**rowcontent)
                    break
                elif Simhash(get_feature(str(buf_list[i]['address_bak']).decode('UTF-8')))\
                        .distance(Simhash(get_feature(u''+rowcontent['address'])))<5:
                    is_correct=0
                    rowcontent['wrong_reason']=u'与历史地址第'+str(buf_list[i]['order_id'])+u'条重复'
                    rowcontent['dup_ID']=buf_list[i]['order_id']
                    rowcontent['dup_ex']=buf_list[i]['ex_id']
                    rowcontent['dup_address']=buf_list[i]['address_bak']
                    rowcontent['dup_name']=buf_list[i]['user_name']
                    rowcontent['dup_phone']=buf_list[i]['mobile']
                    db.pampers_history_dup.insert(**rowcontent)
                    break
                else:continue
        if is_correct==1:
            if rowcontent['user_name'] ==ws1.cell(row=rows-1,column=3).value:
                db.pampers_self_dup.insert(**rowcontent)
            else:
                if (rowcontent['county'].endswith(u'县'))or\
                (rowcontent['county']==u'清新区')or\
                (rowcontent['county']==u'伊宁市')or\
                (rowcontent['county']==u'万山区')or\
                (rowcontent['county']==u'六枝特区'):
                    if((u'村'in rowcontent['address'])and(u'乡村'not in rowcontent['address']))or\
                    ((u'乡'in rowcontent['address'])and(u'乡村'not in rowcontent['address']))or\
                    ((u'镇'in rowcontent['address'])and(u'小镇'not in rowcontent['address'])):
                        rowcontent['class_city']='village'
                    else:
                        rowcontent['class_city']='D'
                elif (rowcontent['county'].endswith(u'市'))and\
                    (rowcontent['county']!=u'巢湖市')and\
                    (rowcontent['county']!=u'毕节市'):
                    rowcontent['class_city']='C'
                elif (rowcontent['city'] in top4_city):
                    rowcontent['class_city']='Top4'
                elif(rowcontent['city'] in A_city):
                    rowcontent['class_city']='A'
                else:
                    rowcontent['class_city']='B'
                db.pampers_order.insert(**rowcontent)

    for rows in range(1,647):
        rowcontent={}
        for cols in range(1,ws1.max_column+1):
            rowcontent[header[cols-1]]=ws1.cell(row=rows,column=cols).value
        is_addr_complete=1
        flag=0
        for prov in provinces:
            if prov in unicode(ws1.cell(row=rows,column=8).value):
                flag=1
                spare_addr0=ws1.cell(row=rows,column=8).value.replace(prov,'')
                spare_addr0=spare_addr0.replace(u'省','')
                if prov in zhixia:
                    rowcontent[header[4]]=prov
                    rowcontent[header[5]]=prov+u'市'
                    spare_addr1=spare_addr0[spare_addr0.find(u'市')+1:]
                elif prov in texing:
                    rowcontent[header[4]]=prov+u'特别行政区'
                elif prov==u'广西':
                    rowcontent[header[4]]=prov+u'壮族自治区'
                    spare_addr1=spare_addr0.replace(u'壮族自治区','')
                elif prov==u'新疆':
                    rowcontent[header[4]]=prov+u'维吾尔自治区'
                    spare_addr1=spare_addr0.replace(u'维吾尔自治区','')
                elif prov==u'宁夏':
                    rowcontent[header[4]]=prov+u'回族自治区'
                    spare_addr1=spare_addr0.replace(u'回族自治区','')
                elif prov==u'西藏':
                    rowcontent[header[4]]=prov+u'自治区'
                    spare_addr1=spare_addr0.replace(u'自治区','')
                elif prov==u'内蒙古':
                    rowcontent[header[4]]=prov+u'自治区'
                    spare_addr1=spare_addr0.replace(u'自治区','')
                else:
                    rowcontent[header[4]]=prov+u'省'
                    spare_addr1=spare_addr0[spare_addr0.find(u'省')+1:]
                break
        if flag==0:
            is_addr_complete=0
            spare_addr1=ws1.cell(row=rows,column=8).value
        if (u'市'in spare_addr1):
            addr2=spare_addr1[0:spare_addr1.find(u'市')+1]
            spare_addr2=spare_addr1[spare_addr1.find(u'市')+1:]
            rowcontent[header[5]]=addr2
        elif u'自治州'in spare_addr1:
            addr2=spare_addr1[0:spare_addr1.find(u'自治州')+3]
            spare_addr2=spare_addr1[spare_addr1.find(u'自治州')+3:]
            rowcontent[header[5]]=addr2
        elif u'地区'in spare_addr1:
            addr2=spare_addr1[0:spare_addr1.find(u'地区')+2]
            spare_addr2=spare_addr1[spare_addr1.find(u'地区')+2:]
            rowcontent[header[5]]=addr2
        elif u'盟'in spare_addr1:
            addr2=spare_addr1[0:spare_addr1.find(u'盟')+1]
            spare_addr2=spare_addr1[spare_addr1.find(u'盟')+1:]
            rowcontent[header[5]]=addr2
        else:
            spare_addr2=spare_addr1
            if rowcontent[header[4]] not in zhixia:
                is_addr_complete=0

        if u'县'in spare_addr2:
            addr3=spare_addr2[0:spare_addr2.find(u'县')+1]
            spare_addr3=spare_addr2[spare_addr2.find(u'县')+1:]
            rowcontent[header[6]]=addr3
        elif (u'市'in spare_addr2)and (u'超市' not in spare_addr2)and (u'市场' not in spare_addr2):
            addr3=spare_addr2[0:spare_addr2.find(u'市')+1]
            spare_addr3=spare_addr2[spare_addr2.find(u'市')+1:]
            rowcontent[header[6]]=addr3
        elif (u'区'in spare_addr2):
            addr3=spare_addr2[0:spare_addr2.find(u'区')+1]
            spare_addr3=spare_addr2[spare_addr2.find(u'区')+1:]
            rowcontent[header[6]]=addr3
        else:
            spare_addr3=spare_addr2
            is_addr_complete=0

        if is_addr_complete==0:
            db.pampers_addr_lack.insert(**rowcontent)
        else:
            buf_list=db((db.history_order.province==ws1.cell(row=rows, column=5).value)&\
                        (db.history_order.city==ws1.cell(row=rows, column=6).value)&\
                        (db.history_order.county==ws1.cell(row=rows, column=7).value)).select\
                (db.history_order.mobile,db.history_order.address_bak,db.history_order.order_id,
                 db.history_order.ex_id,db.history_order.user_name).as_list()
            is_dup=0
            for i in range(0,len(buf_list)):
                if rowcontent['mobile']==buf_list[i]['mobile']:
                    is_dup=1
                    rowcontent['wrong_reason']=u'与历史号码第'+unicode(buf_list[i]['order_id'])+u'条重复'
                    rowcontent['dup_ID']=buf_list[i]['order_id']
                    rowcontent['dup_ex']=buf_list[i]['ex_id']
                    rowcontent['dup_address']=buf_list[i]['address']
                    rowcontent['dup_name']=buf_list[i]['user_name']
                    rowcontent['dup_phone']=buf_list[i]['mobile']
                    db.pampers_history_dup.insert(**rowcontent)
                    break
                elif Simhash(get_feature(buf_list[i]['address'])).distance(Simhash(get_feature(ws1.cell(row=rows,column=10).value)))<24:
                    is_dup=1
                    rowcontent['wrong_reason']=u'与历史地址第'+unicode(buf_list[i]['order_id'])+u'条重复'
                    rowcontent['dup_ID']=buf_list[i]['order_id']
                    rowcontent['dup_ex']=buf_list[i]['ex_id']
                    rowcontent['dup_address']=buf_list[i]['address']
                    rowcontent['dup_name']=buf_list[i]['user_name']
                    rowcontent['dup_phone']=buf_list[i]['mobile']
                    db.pampers_history_dup.insert(**rowcontent)
                    break
                else:continue
            if is_dup==0:
                if rowcontent['user_name']== ws1.cell(row=rows-1, column=3).value:
                    db.pampers_self_dup(**rowcontent)
                else:
                    if (rowcontent['county'].endswith(u'县'))or\
                    (rowcontent['county']==u'清新区')or\
                    (rowcontent['county']==u'伊宁市')or\
                    (rowcontent['county']==u'万山区')or\
                    (rowcontent['county']==u'六枝特区'):
                        if((u'村'in rowcontent['address'])and(u'乡村'not in rowcontent['address']))or\
                        ((u'乡'in rowcontent['address'])and(u'乡村'not in rowcontent['address']))or\
                        ((u'镇'in rowcontent['address'])and(u'小镇'not in rowcontent['address'])):
                            rowcontent['class_city']='village'
                        else:
                            rowcontent['class_city']='D'
                    elif (rowcontent['county'].endswith(u'市'))and\
                        (rowcontent['county']!=u'巢湖市')and\
                        (rowcontent['county']!=u'毕节市'):
                        rowcontent['class_city']='C'
                    elif (rowcontent['city'] in top4_city):
                        rowcontent['class_city']='Top4'
                    elif(rowcontent['city'] in A_city):
                        rowcontent['class_city']='A'
                    else:
                        rowcontent['class_city']='B'
                    db.pampers_order.insert(**rowcontent)

def pampers_census():
    row_content={}
    row_content['tag']='November'
    row_content['distribution']=db(db.pampers_order.order_id >= 0).count()
    row_content['Top4']=db(db.pampers_order.class_city=='Top4').count()
    row_content['A_city']=db(db.pampers_order.class_city=='A').count()
    row_content['B_city']=db(db.pampers_order.class_city=='B').count()
    row_content['C_city']=db(db.pampers_order.class_city=='C').count()
    row_content['D_city']=db(db.pampers_order.class_city=='D').count()
    row_content['village']=db(db.pampers_order.class_city=='village').count()
    db.pampers_result.insert(**row_content)
    row_content={}
    row_content['tag']='percent'
    row_content['distribution']='100%'
    row_content['Top4']=float(db(db.pampers_order.class_city=='Top4').count())/db(db.pampers_order.order_id >= 0).count()
    row_content['A_city']=float(db(db.pampers_order.class_city=='A').count())/db(db.pampers_order.order_id >= 0).count()
    row_content['B_city']=float(db(db.pampers_order.class_city=='B').count())/db(db.pampers_order.order_id >= 0).count()
    row_content['C_city']=float(db(db.pampers_order.class_city=='C').count())/db(db.pampers_order.order_id >= 0).count()
    row_content['D_city']=float(db(db.pampers_order.class_city=='D').count())/db(db.pampers_order.order_id >= 0).count()
    row_content['village']=float(db(db.pampers_order.class_city=='village').count())/db(db.pampers_order.order_id >= 0).count()
    db.pampers_result.insert(**row_content)

def add_PG(ws3):
    product_dict={}
    header=['SB_Internal_Member_ID','Member_Account_ID','Reward_ID','Reward_Name','Reward_Value',
        'Reward_Points','Reward_Internal_Cost','Issue_Date','Issue_Location','Redeemed_Date',
        'Redeemed_Location','Cancel_Date','Expiration_Date','Reward_Barcode','First_Name','Last_Name',
        'Address','City','trade_state','ZipCode','Email_Address','Mobil_Phone','IssuedFlag','RedeemFlag',
        'CanceledFlag','REDEMPTION_STATUS','DLVRYNBR_REJRSN','DELIVERY_ADDRESS']
    for rows in range(1,ws3.max_row+1):
        rowcontent={}
        buf_list=[]
        for cols in range(1,ws3.min_col+1):
            rowcontent[header[cols-1]]=unicode(ws3.cell(row=rows,column=cols).value)
        db.original_data.insert(**rowcontent)
#product id detribute info
    rows= db().select(db.original_data.Reward_Name,
                db.original_data.Reward_Value,
                db.original_data.Reward_Points,
                db.original_data.Reward_Internal_Cost,
                orderby=db.original_data.Reward_Name)
    rows[0]['product_id']=1*100+37
    product_id=1*100+37
    product_dict[rows[0]['Reward_Name']]=rows[0]['product_id']
    db.product_info.insert(**rows[0])
    for i in range(1,len(rows)):
        if (rows[i]['Reward_Name']!=rows[i-1]['Reward_Name'])or\
        (rows[i]['Reward_Value']!=rows[i-1]['Reward_Value'])or\
        (rows[i]['Reward_Points']!=rows[i-1]['Reward_Points'])or\
        (rows[i]['Reward_Internal_Cost']!=rows[i-1]['Reward_Internal_Cost']):
            rows[i]['product_id']=product_id+100
            product_id=rows[i]['product_id']
            product_dict[rows[i]['Reward_Name']]=rows[i]['product_id']
            db.product_info.insert(**rows[i])
        else:pass
#user id detribute info
    user_row={}
    users_address=[]
    address_buf={}
    users= db(db.original_data.DELIVERY_ADDRESS !=None).select\
            (db.original_data.SB_Internal_Member_ID,
             db.original_data.Member_Account_ID,
             db.original_data.DELIVERY_ADDRESS,
             db.original_data.Reward_Name,
             orderby=db.original_data.DELIVERY_ADDRESS)
    users_address=((users[0]['DELIVERY_ADDRESS']).decode('utf8')).split(";")
    user_row['user_name']=users_address[0]
    address_buf['user_name']=users_address[0]
    user_row['address']=users_address[1]
    address_buf['address']=users_address[1]
    user_row['phone_num']=users_address[2]
    address_buf['phone_num']=users_address[2]
    user_row['city']=users_address[3]
    user_row['province']=users_address[4]
    user_row['zip_code']=users_address[5]
    user_row['Country']=users_address[6]
    user_row['SB_Internal_Member_ID']=unicode(users[0]['SB_Internal_Member_ID'])
    user_row['Member_Account_ID']=unicode(users[0]['Member_Account_ID'])
    user_row['user_id']=1*100+79
    user_id=1*100+79
    db.user_info.insert(**user_row)

    for i in range(1,len(users)):
        users_address=((users[i]['DELIVERY_ADDRESS']).decode('utf8')).split(";")
        if len(users_address)==7:
            user_row['user_name']=users_address[0]
            user_row['address']=users_address[1]
            user_row['phone_num']=users_address[2]
            user_row['city']=users_address[3]
            user_row['province']=users_address[4]
            user_row['zip_code']=users_address[5]
            user_row['Country']=users_address[6]
            user_row['SB_Internal_Member_ID']=unicode(users[0]['SB_Internal_Member_ID'])
            user_row['Member_Account_ID']=unicode(users[0]['Member_Account_ID'])
            w1=(user_row['phone_num']==address_buf['phone_num'])
            w2=(Simhash(get_feature(user_row['address'])).distance(Simhash(get_feature(address_buf['address'])))<5)
            w3=(user_row['user_name']==address_buf['user_name'])
            if (w1*0.5+w2*0.3+w3*0.2)<0.5:
                user_row['user_id']=user_id+100
                user_id=user_row['user_id']
                db.user_info.insert(**user_row)
                address_buf['phone_num']=user_row['phone_num']
                address_buf['address']=user_row['address']
                address_buf['user_name']=user_row['user_name']
            else:pass
#user2product
    all_user=db().select(db.user_info.ALL)
    for each in range(0,len(all_user)):
        u2p_row={}
        u2p_row['product_id']=[]
        u2p_row['trials_count']=[]
        u2p_row['user_id']=all_user[each]['user_id']
        u2p_ship=db(int(all_user[each]['Member_Account_ID'])==db.original_data.Member_Account_ID).select\
                (db.original_data.Reward_Name)
        all_product=[]
        all_product.append(u2p_ship[0]['Reward_Name'])
        if len(u2p_ship)>=2:
            for j in range(1,len(u2p_ship)):
                if u2p_ship[j]['Reward_Name'] != u2p_ship[j-1]['Reward_Name']:
                    all_product.append(u2p_ship[j]['Reward_Name'])
        for k in range(0,len(all_product)):
            (u2p_row['product_id']).append(all_product[k])
            each_product_count=all_product.count(all_product[k])
            (u2p_row['trials_count']).append(int(each_product_count))
        db.user2product.insert(**u2p_row)
        del all_product
        del u2p_ship
        u2p_row.clear()

def display_u2p():
    grid = SQLFORM.grid(db.user2product)
    return locals()

def add_babybox_BBS_CD(ws4):
    field_list=['order_id','trade_order','trade_id','ex_id','product_order','barcode',
                    'username','cellphone','province','city','county','address','reserve1',
                    'product_name','data_date','channel','moon','reserve2','product','product_size',
                    'product_brand','store_size','product_quantity','store_name',
                    'receive_province','receive_city','receive_county','receive_address']
    header_BBS=[]
    header_FSCD=[]
    SameOneHouse_phonenum=[]
    town5times_phonenum=[]
    village3times_phonenum=[]
    AddressBlur_phonenum=[]
    cor=[]
    incor=[]
    table_header={}
    for cols in range(1,ws4.min_col+1):
        table_header[field_list[cols-1]]=ws4.cell(row=1,column=cols).value
    for rows in range(2,ws4.max_row+1):
        rowcontent={}
        for cols in range(1,ws4.min_col+1):
            rowcontent[field_list[cols-1]]=ws4.cell(row=rows,column=cols).value
        address=rowcontent['address']
        if address.endswith(u'市')or address.endswith(u'县')or\
        address.endswith(u'镇')or address.endswith(u'工业区')or address.endswith(u'工业园区')or\
        address.endswith(u'公园')or address.endswith(u'管理区')or address.endswith(u'开发区')or\
        address.endswith(u'新区')or address.endswith(u'球场')or address.endswith(u'停车场')or\
        address.endswith(u'篮球场')or address.endswith(u'大道')or address.endswith(u'国道')or\
        address.endswith(u'道')or address.endswith(u'路')or address.endswith(u'桥')or\
        address.endswith(u'路口')or address.endswith(u'街上')or address.endswith(u'街')or\
        address.endswith(u'门口')or address.endswith(u'市场')or address.endswith(u'公交站')or\
        address.endswith(u'车站')or address.endswith(u'加油站')or address.endswith(u'服务站')or\
        address.endswith(u'自取')or address.endswith(u'交叉口')or address.endswith(u'三岔口')or\
        address.endswith(u'桥头')or address.endswith(u'附近')or address.endswith(u'快递')or\
        address.endswith(u'物流')or address.endswith(u'自取')or address.endswith(u'交口'):
            incor.append(rowcontent['cellphone'])
        elif address.endswith(u'号楼') or address.endswith(u'栋')or\
        address.endswith(u'幢') or address.endswith(u'座'):
            SameOneHouse_phonenum.append(rowcontent['cellphone'])
        elif address.endswith(u'小区')or address.endswith(u'校区')or\
        address.endswith(u'宿舍')or address.endswith(u'公寓')or\
        address.endswith(u'期')or address.endswith(u'A区')or\
        address.endswith(u'B区')or address.endswith(u'C区')or\
        address.endswith(u'D区')or address.endswith(u'东区')or\
        address.endswith(u'西区')or address.endswith(u'北区')or\
        address.endswith(u'南区'):
            town5times_phonenum.append(rowcontent['cellphone'])
        elif address.endswith(u'乡')or address.endswith(u'村')or address.endswith(u'庄')or\
        address.endswith(u'营')or address.endswith(u'寨')or address.endswith(u'崖')or\
        address.endswith(u'矿')or address.endswith(u'岗')or address.endswith(u'沟')or\
        address.endswith(u'集')or address.endswith(u'湾')or address.endswith(u'小学')or\
        address.endswith(u'中学')or  address.endswith(u'大学')or address.endswith(u'厂')or\
        address.endswith(u'街道')or address.endswith(u'里')or address.endswith(u'巷')or\
        address.endswith(u'弄')or address.endswith(u'社区')or address.endswith(u'中心')or\
        address.endswith(u'办事处')or address.endswith(u'政府单位')or address.endswith(u'局')or\
        address.endswith(u'法院')or address.endswith(u'医院')or address.endswith(u'大厦')or\
        address.endswith(u'居委会')or address.endswith(u'派出所')or address.endswith(u'汽配城')or\
        address.endswith(u'商贸城')or address.endswith(u'速递易')or address.endswith(u'镇邮局'):
            village3times_phonenum.append(rowcontent['cellphone'])
        elif address.endswith(u'侧') or address.endswith(u'旁')or address.endswith(u'边')or\
        address.endswith(u'隔壁')or address.endswith(u'附近')or address.endswith(u'对面')or \
        address.endswith(u'段'):#or address.endswith(u'门口') or address.endswith(u'口') :
            AddressBlur_phonenum.append(rowcontent['cellphone'])
        else:
            cor.append(rowcontent['cellphone'])
    for rows in range(2,ws4.max_row+1):
        rowcontent={}
        correct_phonenum=SameOneHouse_phonenum + town5times_phonenum + \
                        village3times_phonenum + AddressBlur_phonenum
        phonenum = ws4.cell(row=rows,column=8).value
        for cols in range(1,ws4.min_col+1):
            rowcontent[field_list[cols-1]]=ws4.cell(row=rows,column=cols).value
        if phonenum in incor:
            db.incorrect_BCB.insert(**rowcontent)
        elif phonenum in cor:
            db.correct_BCB.insert(**rowcontent)
        elif phonenum in SameOneHouse_phonenum:
            if SameOneHouse_phonenum.count(phonenum)>=2:
                db.incorrect_BCB.insert(**rowcontent)
            else:
                db.correct_BCB.insert(**rowcontent)
        elif phonenum in town5times_phonenum:
            if (town5times_phonenum.count(phonenum)>5):
                db.incorrect_BCB.insert(**rowcontent)
            else:
                db.correct_BCB.insert(**rowcontent)
        elif phonenum in village3times_phonenum:
            if village3times_phonenum.count(phonenum)>3:
                db.incorrect_BCB.insert(**rowcontent)
            else:
                db.correct_BCB.insert(**rowcontent)
        elif phonenum in AddressBlur_phonenum:
            if AddressBlur_phonenum.count(phonenum)>=2:
                db.incorrect_BCB.insert(**rowcontent)
            else:
                db.correct_BCB.insert(**rowcontent)
        else:pass

def statistics(list1):
    dict1={}
    for i in range(0,len(list1)):
        if  list1[i] in dict1.keys():
            continue
        else:
            li=[]
            li.append(i)
            for j in range(i+1,len(list1)):
                if list1[i] == list1[j]:
                    li.append(j)
        dict1[list1[i]]=li
    return dict1

def add_BBS_20151222(ws4):
    field_list=['order_id','ex_id','shop_name','telephone','cellphone','user_name','address',
                'province','city','county','product_order','barcode','product_name','product_size',
                'store_name','store_size','product_quantity']
    black_name=[]
    full_line=2
    for rows in range(2,ws4.max_row+1):
        if (ws4.cell(row=rows,column=8).value ==u'-'):
            full_line=full_line+1
        else:
            break
#get the all hash Address
    hash_addr=[]
    for i in range(1,ws4.max_row+1):
        hash_addr.append(Simhash(get_feature(ws4.cell(row=i,column=7).value)).value)
#collect the duplicated address_place
    addr_dict={}
    dup_addr_place=[]
    for i in range(0,len(hash_addr)):
        if i in dup_addr_place:
            continue
        else:
            li=[]
            for j in range(i+1,len(hash_addr)):
                if himin_distance(hash_addr[i],hash_addr[j])<8:
                    li.append(j)
            if len(li)==0:
                continue
            else:
                dup_addr_place.extend(li)
                addr_dict[i]=li
#collect the duplicated phone_place
    phone_dict={}
    for i in range(1,ws4.max_row+1):
        if  ws4.cell(row=i,column=5).value in phone_dict.keys():
            continue
        else:
            li=[]
            li.append(i)
            for j in range(i+1,ws4.min_col+1):
                if ws4.cell(row=j,column=5).value == ws4.cell(row=i,column=5).value:
                    li.append(j)
        phone_dict[ws4.cell(row=i,column=5).value]=li
    def InsertExcel2IncorrectDb(i,wrong_reason):
        rowcontent={}
        for cols in range(1,ws4.min_col+1):
            rowcontent[field_list[cols-1]]=ws4.cell(row=i,column=cols).value
        rowcontent['wrong_reason']=wrong_reason
        db.incorrect_BCB.insert(**rowcontent)
    incorrect_record=[]
#process the dup phone_dict
    for phone_place in phone_dict.values():
        if len(phone_place)>=3:
            return u"数据错误"
        else:
            pass
#process the dup addr
    for i,k in addr_dict.items():
        if len(k)>=2:
            k.append(i)
            for j in k:
                InsertExcel2IncorrectDb(j+2,u'当份重复')
                incorrect_record.append(j+2)
        else:
            sizei=str(ws4.cell(row=i+2,column=15).value)
            sizek=str(ws4.cell(row=k[0]+2,column=15).value)
            if (sizei[sizei.find(u'片')-2]==1)and(sizek[sizek.find(u'片')-2]==1):
                InsertExcel2IncorrectDb(k[0]+2,u'当份重复')
                incorrect_record.append(k[0]+2)
            elif (sizei[sizei.find(u'片')-2]!=sizek[sizek.find(u'片')-2]):
                if (ws4.cell(row=i+2,column=5).value)!=(ws4.cell(row=k[0]+2,column=5).value):
                    x = i if sizei[sizei.find(u'片')-2]==1 else k[0]
                    InsertExcel2IncorrectDb(x+2,u'当份重复')
                    incorrect_record.append(x+2)
            else:
                InsertExcel2IncorrectDb(k[0]+2,u'当份重复')
                incorrect_record.append(k[0]+2)
    all_record=range(full_line,ws4.max_row+1)
    precor_record=set(all_record)-set(incorrect_record)

    for r in precor_record:
        address=ws4.cell(row=r,column=7).value
        if address.endswith(u'市')or address.endswith(u'县')or\
        address.endswith(u'镇')or address.endswith(u'工业区')or address.endswith(u'工业园区')or\
        address.endswith(u'公园')or address.endswith(u'管理区')or address.endswith(u'开发区')or\
        address.endswith(u'新区')or address.endswith(u'球场')or address.endswith(u'停车场')or\
        address.endswith(u'篮球场')or address.endswith(u'大道')or address.endswith(u'国道')or\
        address.endswith(u'道')or address.endswith(u'路')or address.endswith(u'桥')or\
        address.endswith(u'路口')or address.endswith(u'街上')or address.endswith(u'街')or\
        address.endswith(u'门口')or address.endswith(u'市场')or address.endswith(u'公交站')or\
        address.endswith(u'车站')or address.endswith(u'加油站')or address.endswith(u'服务站')or\
        address.endswith(u'自取')or address.endswith(u'交叉口')or address.endswith(u'三岔口')or\
        address.endswith(u'桥头')or address.endswith(u'附近')or address.endswith(u'快递')or\
        address.endswith(u'物流')or address.endswith(u'自取')or address.endswith(u'交口'):
            InsertExcel2IncorrectDb(r,u'地址不全')
        else:
            rowcontent={}
            for cols in range(1,ws4.min_col+1):
                rowcontent[field_list[cols-1]]=ws4.cell(row=r,column=cols).value
            address=rowcontent['address']+u''
            address=address.replace(u'，','')
            address=address.replace(u'--',u'-')
            address=address.replace(u'“','')
            address=address.replace(u'”','')
            rowcontent['address']=address
            if u'街道办事处' in address[0:address.find(u'省')]:
                rowcontent['address']=str(rowcontent['address']).replace(u'街道办事处','')
            buf_list=db((db.history.city==ws4.cell(row=r, column=9).value)&\
                        (db.history.county==ws4.cell(row=r, column=10).value)).select\
                    (db.history.order_id,db.history.address_bak,db.history.mobile).as_list()
            is_dup=0
            for i in range(0,len(buf_list)):
                if rowcontent['cellphone']==buf_list[i]['mobile']:
                    is_dup=1
                    rowcontent['wrong_reason']=u'与历史号码第'+unicode(buf_list[i]['order_id'])+u'条重复'
                    db.incorrect_BCB.insert(**rowcontent)
                    break
                elif himin_distance(hash_addr[r-2],Simhash(get_feature(str(buf_list[i]['address_bak']).decode('UTF-8'))).value)<18:
                    is_dup=1
                    rowcontent['wrong_reason']=u'与历史地址第'+unicode(buf_list[i]['order_id'])+u'条重复'
                    db.incorrect_BCB.insert(**rowcontent)
                    break
                else:continue
            if is_dup==0:
                db.correct_BCB.insert(**rowcontent)

def index():
    """
    example action using the internationalization operator T and flash
    rendered by views/default/index.html or views/generic.html
    if you need a simple wiki simply replace the two lines below with:
    return auth.wiki()
    """
    #response.flash = T("Hello World")
    #return dict(message=T('Welcome to web2py!'))
    if request.vars.csvfile1 != None:
        ws=import_excel(request.vars.csvfile1.file)
        add_excel(ws)
        response.flash = T('data uploaded')
    if request.vars.csvfile3 != None:
        ws3=import_excel(request.vars.csvfile3.file)
        add_PG(ws3)
        response.flash = T('data uploaded')
    if request.vars.csvfile4 != None:
        ws4=import_excel(request.vars.csvfile4.file)
        add_BBS_20151222(ws4)
        response.flash = T('data uploaded')
    if request.vars.csvfile2 != None:
        ws1=import_excel(request.vars.csvfile2.file)
        lice=[1516161,1520345,1518691,1542961,1551748,1541116,1516035,1520803,
        1522679,1527299,1538111,1544564,1550749,1520841,1540979,1523641]
        if db(db.pampers_order).isempty():
            add_pampers(ws1)
            response.flash = T('data uploaded')
        else:
            count=0
            for i in range(0,len(lice)):
                if(ws1.cell(row=i,column=1).value==lice[i]):
                    count=count+1
            if count>11:
                add_pampers(ws1)
                response.flash = T('data uploaded')
            else:pass
        pampers_census()
    return dict()

@cache.action()
def download():
    """
    allows downloading of uploaded files
    http://..../[app]/default/download/[filename]
    """
    return response.download(request, db)

def call():
    """
    exposes services. for example:
    http://..../[app]/default/call/jsonrpc
    decorate with @services.jsonrpc the functions to expose
    supports xml, json, xmlrpc, jsonrpc, amfrpc, rss, csv
    """
    return service()


def MD5Sign(md5Str):
    m=hashlib.md5()
    m.update(md5Str)
    return m.hexdigest().upper()
#@auth.requires_login()
@request.restful()
def TradeAdd():
    response.view ='generic.'+request.extension  #return  json
    def POST(tablename,**vars):
        msg=[]
        jsonObj={}
        keys=vars.keys()
        for k in keys:
            k.decode('unicode_escape')
        if ('apiKey' in keys)and('product_info' in keys)and\
        ('timestamp' in keys)and('sig' in keys) and\
        ('out_tid' in keys)and ('shop_id' in keys)and\
        ('consignee' in keys)and('address' in keys)and\
        ('postcode' in keys)and('mobilPhone' in keys)and\
        ('order_date' in keys)and\
        ('product_title' in vars['product_info'][0].keys())and \
        ('standard' in vars['product_info'][0].keys())and\
        ('out__tid' in vars['product_info'][0].keys()):
            ts=int(time.time())
            this_apiKey='A6BEA59B'
            this_apiSecret='1F6F088755B094DDAD3C7AEFEA73A1A1'
            essencial_field=vars.keys()
            essencial_field.sort()
            SigStr=""
            for field in essencial_field:
                if field =="sig" or field =="apiSecret":
                    continue
                elif field =="product_info":
                    SigStr=SigStr+unicode(field)+u"="+u"["
                    for d in vars["product_info"]:
                        k=d.keys()
                        k.sort()
                        SigStr=SigStr+u"{"
                        for j in k:
                            SigStr=SigStr+unicode(j)+u"="+unicode(d[j])+u"&"
                        SigStr=SigStr[0:len(SigStr)-1]
                        SigStr=SigStr+u"}"
                    SigStr=SigStr+u"]"+u"&"
                else:
                     SigStr=SigStr+unicode(field)+u"="+unicode(vars[field])+u"&"
            this_sig=MD5Sign((SigStr[0:len(SigStr)-1]+this_apiSecret).encode('utf8'))
            if (vars['sig']==this_sig) and \
            (ts-int(vars['timestamp'])<180) and \
            ((vars['apiKey']).decode('unicode_escape')==(this_apiKey))  :
                msg.append({'is_success':'true','response_Msg':u'成功导入系统'})
            else:
                msg.append({'sig':(vars['sig']),'this_sig':this_sig,'sigstr':(SigStr[0:len(SigStr)-1]+this_apiSecret).encode('utf8')})
        else:
            msg.append({'is_success':'false','response_Msg':u'absence of essencial field'})
        jsonObj['item']=msg
        return jsonObj
    return locals()

#@auth.requires_login()
@request.restful()
def EdbOrderGet():
    response.view ='generic.'+request.extension  #return  json
    def GET(*args,**vars):
        return int(vars['a'])+int(vars['c'])*int(vars['b'])
    return locals()

def EdbAPI(**arg):
    def _EdbAPI(func):
        def __EdbAPI(data):
            #url="http://vip79.edb04.net/edb2/rest/openapi/index.aspx"
            url="http://vip79.edb04.net/rest/index.aspx"
            SysData={
                "appkey":"5a7b7896",
                "dbhost":"edb_a77527",
                "format":"json",
                "v":2.0,
                "slencry":0,
                "Ip":"117.79.148.228",
                }
            ExData={
                "appscret":"1f5b75edd28d480e968feecbc38f2c73",
                "token":"7041e7424cb4410f8370f10a2d3a285a",
                }
            param=func(data)
            param.update(SysData)
            param["method"]=arg["method"]
            param["timestamp"]=GetTimeStamp()
            paraList=[]
            for k,v in param.items():
                if k=="appkey":
                    continue
                if not str(v):
                    continue
                paraList.append(k+str(v))
            for k,v in ExData.items():
                paraList.append(k+str(v))
            paraList.sort(cmp)
            md5Str="".join(paraList)
            md5Str=SysData["appkey"]+"appkey"+SysData["appkey"]+md5Str
            param["sign"]=MD5Sign(md5Str)
            req=requests.post(url,param)
            if not req.ok:
                print "requests FAIL"
                log_file("%s %s"%(arg,data))
                filePath=log_file(req.content)
                print "log_file in",filePath
                return
            return json.loads(req.content)
        return __EdbAPI
    return _EdbAPI

@EdbAPI(method="edbTradeAdd")
def AddBsgTrade(data):
    infos=data['product_info']
    data.pop('product_info')
    orderXml=XML_ET.fromstring('<info><orderInfo></orderInfo></info>')
    orderInfo=orderXml.find("orderInfo")

    for k,v in data.items():
        tmp=XML_ET.SubElement(orderInfo,k)
        tmp.text=v#.decode('utf8')

    product_info=XML_ET.Element('product_info')
    for info in infos:
        c=XML_ET.SubElement(product_info,'product_item')
        for i,j in info.items():
            d=XML_ET.SubElement(c,i)
            d.text=j

    orderInfo.append(product_info)
    finalStr=XML_ET.tostring(orderXml,"utf8")
    result={}
    result["xmlValues"]=finalStr[37:]
    return result

#@auth.requires_login()
@request.restful()
def TradeGet():
    response.view ='generic.'+request.extension  #return  json
    def GET(*args,**vars):
        patterns=[
            "/products[db_map]",
            "/productid/{db_map.id}",
            "/product_exid/{db_map.field_order}",
            "/product/{custom_order.user_name.startswith}",
            "/product/{custom_order.user_name}/:field",
            "/trades[trade]",
            "/tradeTid/{trade.tid}",
            "/tradeOut/{trade.out_tid}"
        ]
        #patterns = 'auto'
        parser = db.parse_as_rest(patterns,args,vars)
        if parser.status == 200:
            return dict(content=parser.response)
        else:
            raise HTTP(parser.status,parser.error)
    return locals()

@request.restful()
def TradeUpdate():
    response.view ='generic.'+request.extension  #return  json
    def PUT(table_name,record_id,**vars):
        return db(db[table_name].order_id==record_id).update(**vars)
    return locals()

@request.restful()
def TradeDelete():
    response.view ='generic.'+request.extension  #return  json
    def DELETE(table_name,record_id):
        return db(db[table_name].order_id==record_id).delete()
    return locals()




if __name__=="__main__":
    unneed_field=['status','tc_return','edb_return','oti_return','wrong_reason','simid','city_class','addrbkp']
    data0={}
    buf0=db(db.bsg_trade.out_tid=='bsg29').select().as_list()
    buf1=db(db.product_info.out__tid=='bsg29').select().as_list()
    for i,j in buf0[0].items():
        if i in unneed_field:
            continue
        if j:
            if isinstance(j,str):
                data0[i]=j.decode('utf8')
            if isinstance(j, datetime):
                data0[i] = unicode(j)#.isoformat()
            if isinstance(j, long):
                data0[i] = unicode(j)
                #data0[i]=j.decode('utf8')
    info1=[]
    for it in buf1:
        data1={}
        for i,j in it.items():
            if j:
                if isinstance(j,str):
                    data1[i]=j.decode('utf8')
                if isinstance(j, datetime):
                    data1[i] = unicode(j)#.isoformat()
                if isinstance(j, long):
                    data1[i] = unicode(j)
        info1.append(data1)
    data0.update({"product_info":info1})

    edb_return=AddBsgTrade(data0)
    print edb_return





