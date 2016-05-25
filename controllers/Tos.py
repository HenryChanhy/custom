# -*- coding: utf-8 -*-
# 尝试
__author__ = 'Administrator'
from simhash import Simhash,SimhashIndex
import xml.etree.ElementTree as XML_ET
import os
import sys
import time
import hashlib
import requests
import json
import openpyxl
from gluon import DAL,Field
db = DAL(myconf.take('db.uri'), pool_size=1, check_reserved=['all'])

def cmp(x,y):
    x1=x.lower()
    y1=y.lower()
    if x1>y1:
        return 1
    if x1<y1:
        return -1
    return 0

def GetTimeStamp():
    t=time.localtime()
    lst=[]
    for i in xrange(len(t)):
        if i>=5:
            break
        v=str(t[i]).zfill(2)
        lst.append(v)
    reStr="".join(lst)
    return reStr

def log_file(msg):
    dir=os.getcwd()
    fileName="err.txt"
    finalPath=os.path.join(dir,fileName)
    with open(finalPath,"ab") as f:
        f.write("[%s] %s\r\n"%(GetTimeStamp(),msg))
    return finalPath

def MD5Sign(md5Str):
    m=hashlib.md5()
    m.update(md5Str)
    return m.hexdigest().upper()

def TrialAPI(**arg):
    def _TrialAPI(func):
        def __TrialAPI(data):
            SysData={"apiKey":"57Hyjts8HHty",}
            ExData={"apiSecret":"iO7hbmH8-rbt6Hg_Yg6",}
            param=func(data)
            param.update(SysData)
            param.update(data)
            url="http://122.193.31.5:8080/TrialCenter/order/Pampers/ST/"+arg["method"]
            param["timestamp"]=str(int(time.time()))
            paraList=[]
            for k,v in param.items():
                if isinstance(v,str):
                    paraList.append(k+"="+str(v))
            paraList.sort(cmp)
            md5Str="&".join(paraList)
            md5Str=md5Str+ExData["apiSecret"]
            param["sig"]=MD5Sign(md5Str)
            log_file("%s %s"%(md5Str,param["sig"]))
            jsparam=json.dumps(param)
            req=requests.post(url,jsparam,verify=False)
            return json.loads(req.content)       
        return __TrialAPI
    return _TrialAPI

@TrialAPI(method="updateOrderTrackingInfo")
def updateOTI(data):
    return {}

@TrialAPI(method="updateTrialOrderStatus")
def updateTOS(data):
    return {}
def test_TOS():
    wb=openpyxl.load_workbook(r'/home/trade20160125h_36（100条测试71条合格）.xlsx')
    wst= wb.get_sheet_by_name(name = 'true71')
    wsw= wb.get_sheet_by_name(name = 'wrong29')
    for i in range(2,wst.max_row+1):
        outtid=wst.cell(row=i,column=28).value.encode('utf8')
        '''data={"order_id":outtid}
        data["status"]="3"
        result=updateTOS(data)
        log_file("%s %s"%(data,result))'''
        db(db.trade.out_tid==outtid).select().first().update_record(status=3)
    for i in range(2,wsw.max_row+1):
        outtid=wsw.cell(row=i,column=23).value.encode('utf8')
        '''data={"order_id":outtid}
        data["status"]="2"
        data["message"]=wsw.cell(row=i,column=17).value.encode("utf8")
        result=updateTOS(data)
        log_file("%s %s"%(data,result))'''
        db(db.trade.out_tid==outtid).select().first().update_record(status=2)
if __name__=="__main__":
    test_TOS()
