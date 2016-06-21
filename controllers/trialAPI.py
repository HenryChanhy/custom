#!-*- coding:utf-8 -*-

#新接口定义：
#    1、定义一个用TrialAPI装饰的方法，该方法根据需要对传入的参数进行一些必要的处理，并返回一个字典对象
#    2、定义请求结果处理函数，并加入到HandleResult总控方法中
#
#具体接口说明：
#    1、AddTrade:
#        功能:导入订单到e店宝
#        传入参数:必须是以对应的xml标签做key值的字典参数
#    2、GetBarCode:
#        功能：获取产品对应的条形码
#        传入参数:产品名称 或 {"productName":"产品名称"}
#        特殊说明:如果有多条返回值，取第一条返回值中的条形码
__author__ = 'Administrator'
import os
import sys
import time
import hashlib
import chardet
import requests
import xml.etree.ElementTree as XML_ET
import json
#from gluon.contrib.appconfig import AppConfig
import openpyxl
#import response
from gluon import DAL,Field
## once in production, remove reload=True to gain full speed
#myconf = AppConfig(reload=True)
uri	= 'mysql://taotonguser:test578239@localhost/taotongdb'
db = DAL(uri, pool_size=1, check_reserved=['all'])
# -*- coding: utf-8 -*-
#response.generic_patterns = ['*'] if request.is_local else []
#response.formstyle = myconf.take('forms.formstyle')  # or 'bootstrap3_stacked' or 'bootstrap2' or other
#response.form_label_separator = myconf.take('forms.separator')

def GetTimeStamp():
    return str(int(time.time()))

def MD5Sign(md5Str):
    m=hashlib.md5()
    m.update(md5Str)
    return m.hexdigest().upper()

def cmp(x,y):
    x1=x.lower()
    y1=y.lower()
    if x1>y1:
        return 1
    if x1<y1:
        return -1
    return 0

def log_file(msg):
    dir=os.getcwd()
    fileName="err.txt"
    finalPath=os.path.join(dir,fileName)
    with open(finalPath,"ab") as f:
        f.write("[%s] %s\r\n"%(GetTimeStamp(),msg))
    return finalPath

def TrialAPI(**arg):
    def _TrialAPI(func):
        def __TrialAPI(data):
#<<<<<<< HEAD
#=======
#            #url="http://vip802.6x86.com/edb2/rest/index.aspx"
#>>>>>>> 0ce661092826c656b9e7db3a5659a6ab6e6a6caf
            SysData={"apiKey":"57Hyjts8HHty",}
            ExData={"apiSecret":"iO7hbmH8-rbt6Hg_Yg6",}
            param=func(data)
            param.update(SysData)
            param.update(data)
#<<<<<<< HEAD
            url="http://122.193.31.8:8080/TrialCenter/order/Pampers/ST/"+arg["method"]
            param["timestamp"]=str(int(time.time()))
#=======
#            #url="http://IP+PORT/TrialCenter/order/Pampers/ST/"+arg["method"]
#            #url="http://nwct.biz:18910/TrialCenter/order/Pampers/ST/"+arg["method"]
#            url="http://122.193.31.5:8080/TrialCenter/order/Pampers/ST/"+arg["method"]
#            #param["method"]=arg["method"]
#            param["timestamp"]=GetTimeStamp()
#            #param["timestamp"]="201512161115"
#>>>>>>> 0ce661092826c656b9e7db3a5659a6ab6e6a6caf
            paraList=[]
            for k,v in param.items():
                if isinstance(v,str):
                    paraList.append(k+"="+str(v))
            paraList.sort(cmp)
            md5Str="&".join(paraList)
            md5Str=md5Str+ExData["apiSecret"]
            param["sig"]=MD5Sign(md5Str)
            log_file("%s %s"%(md5Str,param["sig"]))
            jsparam=json.dumps(param)
            req=requests.post(url,jsparam,verify=False)
#<<<<<<< HEAD
            return json.loads(req.content)       
#=======
#            return json.loads(req.content)
#>>>>>>> 0ce661092826c656b9e7db3a5659a6ab6e6a6caf
        return __TrialAPI
    return _TrialAPI

@TrialAPI(method="updateOrderTrackingInfo")
def updateOTI(data):
    return {}

@TrialAPI(method="updateTrialOrderStatus")
def updateTOS(data):
    return {}

@TrialAPI(method="TradeAdd")
def addTrade(data):
    return {}

#测试用，可删除
def test_OTI():
    WuLiu_dict={
'EMS':'1','HTKY':'10','ZJS':'11','STO':'12','ZDY':'13','ZDY':'14','YUNDA':'15',
'FEDEX':'16','DBL':'17','RFD':'18','ZT':'19','STO':'2','POSTB':'20','SF':'6',
'STO':'21','OTHER':'22','OTHER':'23','UC':'24','FAST':'25','STO':'26','SF':'7',
'TTKDEX':'27','SF':'28','SF':'29','EMS':'3','ZTO':'30','STO':'31','YUNDA':'32',
'SF':'33','SF':'34','YUNDA':'35','YTO':'36','TTKDEX':'37','YTO':'4','TTKDEX':'5',
'YUNDA':8,'ZTO':9}
    orderinfo={
#<<<<<<< HEAD
	"order_id": "8444",
	"tracking_number": "3100901633843",
	"tracking_company": str(WuLiu_dict['YUNDA'])
#=======
#	"order_id": "4",
#	"tracking_number": "13",
#	"tracking_company": "ZTO"
#>>>>>>> 0ce661092826c656b9e7db3a5659a6ab6e6a6caf
    }
    result=updateOTI(orderinfo)
    log_file("%s %s"%(orderinfo,result))

def test_TOS():
#<<<<<<< HEAD
    wb=openpyxl.load_workbook(r'/home/trade20160125h_36（100条测试71条合格）.xlsx')
    wst= wb.get_sheet_by_name(name = 'true71')
    wsw= wb.get_sheet_by_name(name = 'wrong29')
    for i in range(2,wst.max_row+1):
        outtid=wst.cell(row=i,column=28).value.encode('utf8')
        data={"order_id":outtid}
        data["status"]="3"
        result=updateTOS(data)
        log_file("%s %s"%(data,result))
    db(db.trade.out_tid=='8716').select().first().update_record(status=3)
    for i in range(2,wsw.max_row+1):
        outtid=wsw.cell(row=i,column=23).value.encode('utf8')
        data={"order_id":outtid}
        data["status"]="2"
        data["message"]=wsw.cell(row=i,column=17).value.encode("utf8")
        result=updateTOS(data)
        log_file("%s %s"%(data,result))
        db(db.trade.out_tid==outtid).select().first().update_record(status=2)
    
#=======
#    data={
#    "order_id": "1",
#	"status": "1",
#    "message":"订单重复"
#    }
#    jsonobj= updateTOS(data)
#    return jsonobj
#
#>>>>>>> 0ce661092826c656b9e7db3a5659a6ab6e6a6caf
def test_addTrade():
    data0={"product_totalMoney":"0","storage_id":"11",
       "deliver_status":"\u672a\u53d1\u8d27","consignee":"\u6d4b\u8bd51",
       "actual_RP":"","express":"\u4e2d\u901a\u4e0a\u6d77",
       "pay_status":"\u5df2\u4ed8\u6b3e","invoice_type":"","city":"\u65e0\u9521",
       "is_invoiceOpened":"0","timestamp":"1451903037","buyer_email":"",
       "area":"\u6ee8\u6e56\u533a","province":"\u6c5f\u82cf",
       "process_status":"\u672a\u786e\u8ba4","order_date":"yyyy-mm-dd hh24:mi:ss",
       "mobilPhone":"18651515873","order_totalMoney":"0","buyer_id":"","shop_id":"9",
       "invoice_title":"","out_tid":"1","postcode":"","buyer_alipay":"","pay_method":"",
       "actual_freight_get":"","address":"\u6c5f\u82cf \u65e0\u9521 \u6ee8\u6e56\u533a \u9526\u6eaa\u8def",
       "pay_date":"","invoice_money":"","invoice_msg":"",
       "product_info":[{"orderGoods_Num":"1","out__tid":"1","product_title":"\u6d4b\u8bd5\u5546\u54c11","standard":"1","cost_Price":"0","barCode":"12345678"}],
       "sig":"6b1e5af7b119bdeb1c7dc70512a6be8e","order_type":"\u6b63\u5e38\u8ba2\u5355",
       "finish_date":""}
    for k,v in data0.items():
        if isinstance(v,str):
            data0[k]=v.decode("unicode-escape")
    print addTrade(data0)
if __name__=="__main__":
    test_TOS()
    #test_OTI()
